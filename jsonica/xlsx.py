# -*- coding: utf-8 -*-
# this made for python3
import os, csv, openpyxl
from jsonica import errorout, PROGNAME
from schema_helper import Schema, TypeSign, Validator
from sub_command_core.generate import output_formats
from util import Util, Hoare

class XLSX:
  """ xlsx 具象操作クラス """
  __DEBUG = True
  DEBUG = not (os.getenv('TRAVIS', not __DEBUG))

  # childへのリンクを示す接頭辞
  sheet_link_sign = 'sheet://'

  @property
  def schema(self): return self.__schema
  @schema.setter
  def schema(self, value):
    if not hasattr(self, '__schema'):
      self.__schema = value

  @property
  def piled_schema(self):
    """
    __all_schema: {sheet: [schemas]}
    """
    return self.__all_schema
  @piled_schema.setter
  def piled_schema(self, value):
    """
    schema用accumrator

    :param tuple value: (sheet:Some, column:Some, schema:dict)
    """
    sheet, col, schema = value
    synth_key = '{}/{}'.format(sheet, col)
    if sheet in self.piled_schema.keys():
      self.piled_schema[synth_key].update(schema)
    else:
      self.piled_schema[synth_key] = schema

  def __init__(self, file, enc, root_name=None, forms=None):
    """

    :param str file: xlsx file location

    :param str enc: encodig指定があった場合 default 'utf8'

    :param str root_name: root itemを示すシート名 default 'root'

    :param tuple forms: ?sv 記述フォーマット (?sv, delimiter, output_root_path)
    """
    self.filepath = file
    self.filename = os.path.basename(file)
    self.root_name = root_name
    self.__all_schema = {}
    if forms:
      self.format = forms[0]
      self.format_delimiter = forms[1]
      self.format_output = forms[2]
    self.char_encode = enc
    if forms: # generate
      self.book = openpyxl.load_workbook(self.filepath, keep_vba=True, data_only=False)
    else: # init
      self.book = openpyxl.Workbook()

  def generate_json(self, sheet_name=None, acc=None):
    """
    sheet_nameが指すsheetのJSONをaccに追加する

    :param str sheet_name: 別シートに分かれるitemのシート名 Noneの場合はrootと認識

    :param acc: rootから伝播されるaccumrator

    :return この処理から得られた連想配列が追加されたaccumratorを返す
    """
    sheets = self.__name_to_sheets()
    if not sheet_name:
      sheet_name = self.root_name
    # pyxl...Workbookで[sheet名]を持っているが、あまり高速処理向けではないため
    sheet_names = list(sheets.keys())
    Util.sprint('in process %s'%sheet_name, self.DEBUG)
    Hoare.P(sheet_name in sheet_names, '"%s" not found in %s'%(sheet_name ,sheet_names))
    root_sheet = sheets[sheet_name]
    self.check_charcode(root_sheet)
    columns = []
    acc = [] if not acc else acc
    Util.sprint('I\'ll update {}'.format(acc), self.DEBUG)
    # COMBAK: 処理速度に問題が出るようであれば分散処理検討
    # A1, B1...で場所を特定するか、indexで回すか
    for i, row in enumerate(root_sheet.iter_rows()):
      subacc = {}
      if self.format:
        self.__output_to_csv(self.format_output, root_sheet, self.char_encode)
      for j, cell in enumerate(row):
        v = cell.value # off-by-oneを気にしないといけなくなるので、col_idxではなくenumerate使う
        if v is None: continue # cell check
        if i == 0: # 初行 column check
          # cell.commentは必ずつくが、中身がない場合はNone
          if hasattr(cell, "comment") and cell.comment:
            # column 準備 / schemaは遅延せずこの時点で辞書として成立している事を保証
            columns.append((v, Util.runtime_type(cell.comment.text)))
          else:
            self.errorout(2, 'sheet = {}, col = {}, row = {}'.format(sheet_name, j, i))
        else:
          # TODO: 関数へ置き換え type = array, objectのケース をカバー
          # 別sheet評価
          if isinstance(v, str) and v.startswith(XLSX.sheet_link_sign):
            # COMBAK: sheetであることがarray, objectの必要条件になってしまっている
            # primitive配列をどう表現するかによって改修が必要 __storeに包含させる？
            link = v.lstrip(XLSX.sheet_link_sign)
            if link in sheet_names:
              col_name, col_schema = columns[j]
              Util.sprint('process %s -> %s'%(col_name, link), self.DEBUG)
              Util.sprint('current acc = %s'%acc, self.DEBUG)
              # recursive seed
              XLSX.__store(
                self.generate_leaf(root_sheet.title, col_name, link, col_schema),
                subacc)
            else:
              errorout(1, 'sheet = from %s to %s, col = %d, row = %d'%(sheet_name, link, j, i))
          else:
            XLSX.__store(self.type_validator(sheet_name, v, columns[j]), subacc)
        # pass columns
      Util.check_emptyOR(lambda x: XLSX.__store(x, acc), subacc)
      # pass a row
    return acc

  def __output_to_csv(self, base_path, sheet, enc):
    """
    CSV, TSV出力
    """
    if not self.format: return
    Hoare.P(self.format in output_formats)
    xdest = os.path.join(base_path, self.filename)
    os.makedirs(xdest, exist_ok=True)
    xdest_path = os.path.join(xdest, '%s.%s'%(sheet.title ,self.format))
    Util.sprint(' > %s'%xdest_path, self.DEBUG)
    with open(xdest_path, 'w', encoding=enc) as f:
      writer = csv.writer(f, delimiter=self.format_delimiter)
      for cols in sheet.rows:
        writer.writerow([str(col.value or '') for col in cols])

  def __name_to_sheets(self):
    """
    | sheetを{sheet名:sheet}形式にして返す
    | instance 生成後、実行中のExcel更新は考えない
    """
    # TODO: pyxl.get_sheet_names(), get_sheet_by_name()で代用できるか検討
    if not hasattr(self, '__sheets_cache'):
      self.__sheets_cache = {s.title: s for s in self.book.worksheets}
    return self.__sheets_cache

  def check_charcode(self, item, valid_enc=None):
    Hoare.P(isinstance(item, openpyxl.workbook.workbook.Workbook) or \
            isinstance(item, openpyxl.worksheet.Worksheet) or \
            isinstance(item, openpyxl.cell.Cell))
    enc = valid_enc if bool(valid_enc) else self.char_encode
    Util.sprint(enc, self.DEBUG)
    if not (item.encoding == enc):
      # TODO: sheet, cellごとにエラーを上げる場合の処理
      if isinstance(item, openpyxl.workbook.workbook.Workbook):
        add = 'sheet_names = %s'%item.sheet_names
      elif isinstance(item, openpyxl.worksheet.Worksheet):
        add = 'sheet_name = %s'%item.title
      else: # Cell
        add = 'parent = {} index = {}'.format(item.parent ,item.cordinate)
      print('*'*50)
      print(add)
      print('*'*50)
      # TODO: excel rw 状態チェック
      item.encoding = enc

  def type_validator(self, sheet_name, value, type_desc, validator=Validator.jsonschema):
    """
    | Validator switcher
    | validation を passしたら成功した**評価値のみ**を返す
    | 失敗したら、その場でcommand errorとする

    :param str sheet_name: 評価対象sheet名

    :param value: 評価対象JSON

    :param type_desc: schema (format TBD)
    """
    self.schema = Schema(validator)
    raw = Util.conv_escapedKV(XLSX.__get_type(type_desc[1]), type_desc[0], value)
    instance = Util.runtime_type('{%s}'%raw)
    Util.sprint('i\'m %s. call validator'%self, self.DEBUG)
    self.piled_schema = (sheet_name, type_desc[0], {type_desc[0]:type_desc[1]})
    Util.sprint('>> %s -> type: %s\n%s'%(sheet_name, type_desc, instance), self.DEBUG)
    Hoare.P(instance is not None)
    self.schema.validate(instance, type_desc)
    return instance

  def generate_sheet(self, name):
    return self.book.create_sheet(name)

  def generate_leaf(self, parent, key, l, schema):
    """ schemaに従ったitemを生成 """
    # NOTE: recursive procが分解されている事に留意
    self.piled_schema = (parent, key, schema)
    return {key: self.generate_json(l, XLSX.renew_acc(schema))}

  @classmethod
  def put_cell_comment(cls, cell, text, author=PROGNAME):
    from openpyxl.comments import Comment
    cell.comment = Comment(text, author)

  @classmethod
  def __get_type(cls, schema):
    Hoare.P('type' in schema.keys())
    return schema['type']

  @classmethod
  def renew_acc(cls, schema):
    Hoare.P(isinstance(schema, dict))
    _type = XLSX.__get_type(schema)
    if _type == TypeSign.ARRAY:
      return []
    elif _type == TypeSign.OBJ:
      return {}
    else:
      errorout(4, _type)

  @classmethod
  def __store(cls, item, accumulator):
    """
    評価済みが保証された値を、rootのleafに連結

    :param accumlator: either dict or list
    """
    if isinstance(accumulator, dict):
      accumulator.update(item)
    elif isinstance(accumulator, list):
      accumulator.append(item)
    else:
      errorout(5)
    return accumulator
