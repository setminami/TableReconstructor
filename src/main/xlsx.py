# -*- coding: utf-8 -*-
# this made for python3
import os, sys
import csv, openpyxl
from .jsonica import Jsonica, errorout, PROGNAME
from .schema_helper import Schema, TypeSign, Validator

# from .util import Util

class XLSX:
  """ xlsx 具象操作クラス """
  # childへのリンクを示す接頭辞
  sheet_link_sign = 'sheet://'

  def __init__(self, file, enc, forms=None):
    self.filepath = file
    self.filename = os.path.basename(file)
    if forms:
      self.format = forms[0]
      self.format_delimiter = forms[1]
      self.format_output = forms[2]
    self.char_encode = enc
    if forms: # generate
      self.book = openpyxl.load_workbook(self.filepath, keep_vba=True, data_only=False)
    else: # init
      self.book = openpyxl.Workbook()
    pass

  def generateJSON(self, sheet_name, acc=[]):
    """
    sheet_nameが指すsheetのJSONをaccに追加する
    """
    sheets = self.__nameToSheets()
    # pyxl...Workbookで[sheet名]を持っているが、あまり高速処理向けではないため
    sheet_names = list(sheets.keys())
    self.__print(f'in process {sheet_name}')
    assert sheet_name in sheet_names, f'"{sheet_name}" not found in {sheet_names}'
    root_sheet = sheets[sheet_name]
    self.checkCharEncode(root_sheet)
    columns = []
    self.__print(f'I\'ll update {acc}')
    # Memo: 処理速度に問題が出るようであれば分散処理検討
    # A1, B1...で場所を特定するか、indexで回すか
    for i, row in enumerate(root_sheet.iter_rows()):
      subacc = {}
      if self.format:
        self.__outputCSV(self.format_output, root_sheet, self.char_encode)
        pass
      for j, cell in enumerate(row):
        v = cell.value # off-by-oneを気にしないといけなくなるので、col_idxではなくenumerate使う
        if v is None: continue # cell check
        if i == 0: # 初行 column check
          # cell.commentは必ずつくが、中身がない場合はNone
          if hasattr(cell, "comment") and cell.comment:
            # column 準備 / schemaは遅延せずこの時点で辞書として成立している事を保証
            columns.append((v, Util.runtimeDictionary(cell.comment.text)))
          else:
            self.errorout(2, f'sheet = {sheet_name}, col = {j}, row = {i}')
        else:
          # ToDo: 関数へ置き換え type = array, objectのケース をカバー
          if isinstance(v, str) and v.startswith(XLSX.sheet_link_sign):
            # Memo: sheetであることがarray, objectの必要条件になってしまっている
            # primitive配列をどう表現するかによって改修が必要 __storeに包含させる？
            link = v.lstrip(XLSX.sheet_link_sign)
            if link in sheet_names:
              col_name = columns[j][0]
              self.__print(f'process {col_name} -> {link}')
              self.__print(f'current acc = {acc}')
              new_acc = self.__brandnewAccForType(columns[j][1])
              self.__store({col_name:self.generateJSON(sheet_name=link, acc=new_acc)}, subacc)
            else:
              self.errorout(1, f'sheet = from {sheet_name} to {link}, col = {j}, row = {i}')
              pass
          else:
            self.__store(self.typeValidator(v, columns[j]), accumulator=subacc)
        pass # pass columns
      Util.checkEmptyOr(lambda x: self.__store(x, acc), subacc)
      pass # pass a row
    return acc

  def __getType(self, schema):
    assert 'type' in schema.keys()
    return schema['type']

  def __brandnewAccForType(self, schema):
    assert isinstance(schema, dict)
    _type = self.__getType(schema)
    if _type == TypeSign.ARRAY:
      return []
    elif _type == TypeSign.OBJ:
      return {}
    else:
      errorout(4, _type)

  def __store(self, item, accumulator):
    if isinstance(accumulator, dict):
      accumulator.update(item)
    elif isinstance(accumulator, list):
      accumulator.append(item)
    else:
      errorout(5)
    return accumulator

  def __outputCSV(self, base_path, sheet, enc):
    """
    CSV, TSV出力
    """
    if not self.format: return
    assert self.format in output_formats
    xdest = os.path.join(base_path, self.filename)
    os.makedirs(xdest, exist_ok=True)
    xdest_path = os.path.join(xdest, f'{sheet.title}.{self.format}')
    self.__print(f' > {xdest_path}')
    with open(xdest_path, 'w', encoding=enc) as f:
      writer = csv.writer(f, delimiter=self.format_delimiter)
      for cols in sheet.rows:
        writer.writerow([str(col.value or '') for col in cols])

  def __nameToSheets(self):
    """
    sheetを{sheet名:sheet}形式にして返す
    instance 生成後、実行中のExcel更新は考えない
    """
    # ToDo: get_sheet_names(), get_sheet_by_by_name()で代用できるか検討
    if not hasattr(self, '__sheets_cache'):
      self.__sheets_cache = {s.title: s for s in self.book.worksheets}
    return self.__sheets_cache

  def checkCharEncode(self, item, valid_enc=None):
    assert isinstance(item, openpyxl.workbook.workbook.Workbook) or \
            isinstance(item, openpyxl.worksheet.Worksheet) or \
            isinstance(item, openpyxl.cell.Cell)
    enc = valid_enc if bool(valid_enc) else self.char_encode
    self.__print(enc)
    if not (item.encoding == enc):
      # ToDo: sheet, cellごとにエラーを上げる場合の処理
      if isinstance(item, openpyxl.workbook.workbook.Workbook):
        add = f'sheet_names = {item.sheet_names}'
      elif isinstance(item, openpyxl.worksheet.Worksheet):
        add = f'sheet_name = {item.title}'
      else: # Cell
        add = f'parent = {item.parent} index = {item.cordinate}'
      print('*'*50)
      print(add)
      print('*'*50)
      # ToDo: excel rw
      item.encoding = enc
      pass
    pass

  def typeValidator(self, value, type_desc, validator=Validator.jsonschema):
    """ Validator switch """
    if not hasattr(self, '__schema'):
      print(f'new!! {type_desc}')
      self.__schema = Schema(validator)
    raw = Util.convEscapedKV(self.__getType(type_desc[1]), type_desc[0], value)
    instance = Util.runtimeDictionary('{%s}'%raw)
    self.__schema.validate(instance, type_desc)
    assert instance is not None
    return instance

  def generateSheet(self, name):
    return self.book.create_sheet(name)

  def putCommentToCell(self, cell, text, author=PROGNAME):
    from openpyxl.comments import Comment
    cell.comment = Comment(text, author)

  # SP_FILE 注意
  def __print(self, str, flag=False):
    if flag:
      print(str)
    pass
